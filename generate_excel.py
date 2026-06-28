# -*- coding: utf-8 -*-
"""
v4 - Generador consolidado premium adaptado a la estructura de la demo "Gestion Material Scouts.xlsx".
Estructura de la hoja principal:
MATERIAL | CATEGORÍA | TOTAL (=SUM(Castores:Rutas)) | CASTORES | LOBATOS | EXPLORADORES (TROPA) | PIONEROS | RUTAS | TIPO | DETALLES | CAJA | COMPRADO
Reutiliza los datos curados de data1.py y data2.py, y la lista MASTER de data3.py.
Genera fórmulas de Excel dinámicas vinculadas a la hoja CONFIG (niños por unidad).
Separa la alimentación y los talleres de cocina en la hoja ALIMENTACIÓN.
"""

import sys
import os
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Asegurar path para importar data1, data2, data3
BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BASE)

from data1 import RAW
from data2 import RAW2
try:
    from data3 import MASTER
except ImportError:
    MASTER = []

# Unificar datos curados
ALL_RAW_DATA = {**RAW, **RAW2}
UNITS_ORDER = ["CASTORES", "LOBATOS", "EXPLORADORES", "PIONEROS", "RUTAS PANDORA", "RUTAS ICARO", "COMISIONES"]

# Mapeos de categorías y tipos de data3 a los términos del Excel
CAT_MAP = {
    "Construcción y herramientas": "Construccion/herramientas",
    "Papelería y manualidades": "Papeleria/manualidades",
    "Juegos y deporte": "Juegos/deporte",
    "Cocina y menaje": "Cocina/menaje",
    "Iluminación y electrónica": "Iluminacion/electronica",
    "Higiene, salud y limpieza": "Higiene/limpieza",
    "Textil y disfraces": "Textil/disfraces",
    "Impresiones y documentación": "Impresiones/documentacion",
    "Equipo de campamento": "Equipo campamento",
    "Otros": "Otros"
}

TIPO_MAP = {
    "COMPRAR": "COMPRAR",
    "YA HAY (grupo/bases) - revisar": "YA HAY (bases) - revisar",
    "LO TRAE EL NIÑO": "LO TRAE EL NINO",
    "UNIDAD/COMISIÓN": "UNIDAD/COMISION",
    "PEDIR A COCINA/MENÚ": "PEDIR A COCINA/MENU"
}

# Reglas semánticas para materiales no mapeados en data3
RULES = [
    # Cocina / Menaje (no alimentos)
    (["sarten", "sartén", "sartenes", "cazuela", "cazuelas", "olla", "ollas", "cazo", "espatula", "espátula", 
      "bol", "boles", "mortero", "morteros", "tenedor", "tenedores", "cuchara", "cucharas", "cuchillo", 
      "cuchillos", "pelador", "peladores", "tabla de cortar", "tablas de cortar", "varillas", "batidor", 
      "bandejas", "cazo para servir", "plato", "platos", "vaso", "vasos", "poto", "potos", "camping gas", 
      "hornillo", "hornillos", "quemador", "bombona", "bombonas"], 
     "Menaje de cocina (sartenes, cazuelas, boles, hornillos, bombonas)", "Cocina/menaje", "YA HAY (bases) - revisar"),
    
    # Electrónica e iluminación
    (["altavoz", "altavoces", "reproductor"], "Altavoz bluetooth", "Iluminacion/electronica", "YA HAY (bases) - revisar"),
    (["walkie", "walkies", "walkie talkie", "walkie talkies"], "Walkie-talkies", "Iluminacion/electronica", "YA HAY (bases) - revisar"),
    (["puntero laser", "puntero láser", "láser", "laser"], "Puntero láser", "Iluminacion/electronica", "YA HAY (bases) - revisar"),
    (["frontal", "frontales", "linterna", "linternas"], "Linternas / Frontales", "Iluminacion/electronica", "LO TRAE EL NINO"),
    (["vela", "velas"], "Velas", "Iluminacion/electronica", "COMPRAR"),
    (["enchufe", "alargo", "cable", "pilas", "batería", "baterias"], "Pilas y electricidad", "Iluminacion/electronica", "COMPRAR"),
    
    # Juegos y deportes
    (["balon", "balón", "balones", "pelota", "pelotas", "set de petanca", "petanca"], "Balones y pelotas deportivas / Petanca", "Juegos/deporte", "YA HAY (bases) - revisar"),
    (["cono", "conos"], "Conos deportivos / Conos chinos", "Juegos/deporte", "YA HAY (bases) - revisar"),
    (["aro", "aros"], "Aros de plástico", "Juegos/deporte", "YA HAY (bases) - revisar"),
    (["silbato", "silbatos"], "Silbato", "Juegos/deporte", "YA HAY (bases) - revisar"),
    (["juego de mesa", "juegos de mesa", "cartas", "baraja", "barajas", "trivial", "dados", "fichas", 
      "gobbit", "jungle speed", "virus", "stratego", "estratego", "monopoly"], 
     "Juegos de mesa y barajas de cartas", "Juegos/deporte", "YA HAY (bases) - revisar"),
    (["discos para lanzar", "frisbee", "jugger", "lanzas", "escudos", "espadas", "comba", "combas", "cuerda sogatira"], 
     "Material de juegos especiales (Jugger, frisbee, combas, sogatira)", "Juegos/deporte", "YA HAY (bases) - revisar"),
    
    # Higiene, salud y limpieza
    (["bolsa de basura", "bolsas de basura", "bolsa basura"], "Bolsas de basura (grandes/medianas)", "Higiene/limpieza", "COMPRAR"),
    (["fairy", "estropajo", "estropajos", "detergente", "jabón líquido", "jabón", "jabon", "limpieza"], "Productos de limpieza (Fairy, estropajo, jabón)", "Higiene/limpieza", "COMPRAR"),
    (["papel higienico", "papel higiénico", "papel de cocina", "servilletas"], "Papel higiénico y de cocina", "Higiene/limpieza", "COMPRAR"),
    (["botiquin", "botiquín", "tiritas", "vendas", "venda", "gasa", "gasas", "desinfectante", "tarjetas sanitarias"], 
     "Botiquín de primeros auxilios / Tarjetas sanitarias", "Higiene/limpieza", "YA HAY (bases) - revisar"),
    (["crema solar", "protector solar", "aftersun", "crema hidratante"], "Crema de protección solar / Aftersun", "Higiene/limpieza", "LO TRAE EL NINO"),
    
    # Impresiones y documentación
    (["mapa", "mapas", "brújula", "brújulas", "brujula", "brujulas"], "Mapas de la zona y Brújulas", "Impresiones/documentacion", "YA HAY (bases) - revisar"),
    (["fotocopia", "fotocopias", "impreso", "impresos", "impresas", "imprimir", "leyendas", "cuaderno de caza", 
      "cuadernos", "fichas", "títulos", "titulos", "reflexiones", "roles", "noticias"], 
     "Fotocopias, fichas de actividades y reflexiones", "Impresiones/documentacion", "COMPRAR"),
    
    # Equipo de campamento
    (["tienda", "tiendas", "piquetas", "piqueta", "mástil", "mastil", "lona", "lonas", "sombrajo", "sombrajos", 
      "rafia", "rafias", "picas", "pica"], 
     "Equipo de campamento (Tiendas, lonas, piquetas, rafias, picas)", "Equipo campamento", "YA HAY (bases) - revisar"),
    
    # Textil y disfraces
    (["camiseta", "camisetas", "tote bag", "tote bags", "bolso", "bolsos"], "Textiles para talleres (camisetas, tote bags)", "Textil/disfraces", "LO TRAE EL NINO"),
    (["telas", "tela", "disfraz", "disfraces", "velos", "maquillaje", "pinturas de cara", "pintura cara", "pintacaras"], 
     "Telas, disfraces y maquillaje de caracterización", "Textil/disfraces", "YA HAY (bases) - revisar"),
    
    # Equipo personal (lo trae el niño)
    (["cantimplora", "cantimploras", "mochila", "mochilas", "saco", "sacos", "aislante", "aislantes", "esterilla", 
      "esterillas", "calzado", "bañador", "bañadores", "toalla", "toallas", "chanclas", "gorra", "gorras", 
      "gafas de sol", "ropa", "baño", "natación", "peine"], 
     "Equipo personal (mochila, saco, esterilla, toalla, etc.)", "Otros", "LO TRAE EL NINO"),
    (["guantes"], "Guantes de trabajo (protección)", "Higiene/limpieza", "LO TRAE EL NINO"),
]

def is_food_item(text):
    t = text.lower().strip()
    utensils = {
        "sartén", "sarten", "cazuela", "olla", "mortero", "hornillo", "camping", "camping gas", "gas", 
        "varilla", "varillas", "tenedor", "cuchara", "cuchillo", "plato", "vaso", "cazo", "espatula", 
        "espátula", "bol", "bol grande", "tabla", "tablas", "pelador", "peladores", "vendas", "venda"
    }
    for u in utensils:
        if u in t:
            return False
            
    food_keywords = [
        "ingredientes", "toppings", "compra (20 niños)", "menú", "menu", "comida",
        "harina", "leche", "huevo", "huevos", "azúcar", "azucar", "mantequilla", "sal", "levadura",
        "coco rallado", "leche condensada", "granadina", "cardamomo", "jamón york", "jamón", "queso", "pan",
        "gominolas", "nutella", "nocilla", "chocolate", "mermelada", "gusanitos", "patatas", "patata",
        "cebolla", "tomate frito", "bacon", "pepperoni", "salami", "salchichas", "atún", "anchoas",
        "champiñones", "pimientos", "aceitunas", "maíz", "piña", "rúcula", "mozzarella", "miel",
        "orégano", "pimienta", "ajo", "galletas", "queso crema", "caldo de pollo", "pechugas", "jengibre",
        "cebolleta", "puerros", "salmón", "carne picada", "fruta", "bebidas", "coca-cola", "fanta", "aquarius",
        "manzanas", "lomo", "tranchetes", "paté", "noodles", "salchichón", "sopa", "tang", "frutos secos",
        "masa de empanadillas", "alimentos", "comida del día", "comida volante", "plátanos", "yogures", 
        "vinagre", "ketchup", "mayonesa", "especias"
    ]
    return any(kw in t for kw in food_keywords)

def extract_qty_number(text, material_name):
    """Extrae la cantidad numérica aproximada de la descripción del material."""
    t = text.lower().strip()
    
    # 1. Quitar referencias de tamaño que puedan confundir al buscador
    t = re.sub(r'16\s*mm', '', t)
    t = re.sub(r'10\s*x\s*10\s*cm', '', t)
    
    # 2. Buscar patrones de cantidad ("x2", "x 4", "100m", "4 botes", "2 rollos")
    match = re.search(r'(?:x\s*|\b)(\d+(?:\.\d+)?)\s*(?:rollos|botes|metros|latas|paquetes|rulos|sobres|láminas|laminas|kilos|kg|l|unidades|uds|ud|m\b)', t)
    if match:
        return float(match.group(1))
        
    # 3. Buscar número al principio (ej: "2 barreños", "3 mantas")
    match_start = re.match(r'^(\d+(?:\.\d+)?)\b', t)
    if match_start:
        return float(match_start.group(1))
        
    # 4. Buscar número al final (ej: "imperdibles x4")
    match_end = re.search(r'\bx\s*(\d+(?:\.\d+)?)$', t)
    if match_end:
        return float(match_end.group(1))
        
    if "par de" in t:
        return 2.0
        
    return 1.0 # Asumir 1 si se menciona sin número

def find_in_master(clean_text):
    for item in MASTER:
        name_lower = item[0].lower().strip()
        if clean_text == name_lower:
            return item
    for item in MASTER:
        name_lower = item[0].lower().strip()
        if len(name_lower) > 3:
            name_sing = name_lower.rstrip('s')
            text_sing = clean_text.rstrip('s')
            if name_sing in text_sing or text_sing in name_sing:
                return item
    return None

def normalize_material(item_text, unit, title, date):
    clean_text = item_text.strip()
    
    # Comprobar si es comida
    if is_food_item(clean_text):
        return None
        
    # Buscar en MASTER de data3
    master_match = find_in_master(clean_text.lower())
    if master_match:
        return {
            "name": master_match[0],
            "category": CAT_MAP.get(master_match[1], "Otros"),
            "type": TIPO_MAP.get(master_match[2], "YA HAY (bases) - revisar"),
            "formula_raw": master_match[3],
            "literal": f"{unit} - {title} ({date}): «{item_text}»" if date else f"{unit} - {title}: «{item_text}»",
            "is_master": True
        }
        
    # Buscar en RULES semánticas
    clean_lower = clean_text.lower()
    for kw_list, std_name, cat, tipo in RULES:
        for kw in kw_list:
            if kw in clean_lower:
                return {
                    "name": std_name,
                    "category": cat,
                    "type": tipo,
                    "formula_raw": "",
                    "literal": f"{unit} - {title} ({date}): «{item_text}»" if date else f"{unit} - {title}: «{item_text}»",
                    "is_master": False
                }
                
    # Por defecto
    is_craft = any(x in clean_lower for x in ["papel", "cartulina", "pintura", "pegamento", "cinta", "témpera", "tijeras", "pincel"])
    cat = "Papeleria/manualidades" if is_craft else "Otros"
    tipo = "COMPRAR" if is_craft else "YA HAY (bases) - revisar"
    name_norm = clean_text[0].upper() + clean_text[1:] if clean_text else clean_text
    
    return {
        "name": name_norm,
        "category": cat,
        "type": tipo,
        "formula_raw": "",
        "literal": f"{unit} - {title} ({date}): «{item_text}»" if date else f"{unit} - {title}: «{item_text}»",
        "is_master": False
    }

# ══════════════════════════════════════════════════════════════════════
# PROCESAMIENTO
# ══════════════════════════════════════════════════════════════════════
print("Consolidando datos curados en columnas desglosadas por unidad...")

# Diccionario para almacenar el material consolidado
# Estructura: name -> { category, type, Castores, Lobatos, Tropa, Pioneros, Rutas, detalles }
consolidated = defaultdict(lambda: {
    "category": "", "type": "", 
    "CASTORES": 0, "LOBATOS": 0, "EXPLORADORES": 0, "PIONEROS": 0, "RUTAS": 0, "COMISIONES": 0,
    "detalles": []
})

food_list = [] 
menu_items = [] 

for unit in UNITS_ORDER:
    activities = ALL_RAW_DATA.get(unit, [])
    for date, title, material_text in activities:
        # Menú Volante
        if "MENÚ VOLANTE" in title.upper() or "MENU VOLANTE" in title.upper():
            parts = [p.strip() for p in material_text.split(";") if p.strip()]
            for p in parts:
                menu_items.append((unit, date, title, p))
            continue
            
        parts = [p.strip() for p in material_text.split(";") if p.strip()]
        is_food_act = any(kw in title.lower() for kw in ["taller de cocina", "masterchef", "tortitas", "gyozas", "empanadillas", "tortilla de patata", "picoteo"])
        
        if is_food_act:
            food_parts = []
            for p in parts:
                norm = normalize_material(p, unit, title, date)
                if norm is None:
                    food_parts.append(p)
                else:
                    name = norm["name"]
                    qty_num = extract_qty_number(p, name)
                    
                    # Identificar la columna de destino
                    col_key = "RUTAS" if "RUTAS" in unit else ("EXPLORADORES" if unit == "EXPLORADORES" else ("COMISIONES" if unit == "COMISIONES" else unit))
                    
                    consolidated[name]["detalles"].append(norm["literal"])
                    if not consolidated[name]["category"]:
                        consolidated[name]["category"] = norm["category"]
                        consolidated[name]["type"] = norm["type"]
                    
                    # Guardar cantidad/fórmula
                    consolidated[name][col_key] = qty_num
            if food_parts:
                food_list.append((unit, date, title, "; ".join(food_parts)))
            continue
            
        # Actividad normal
        for p in parts:
            norm = normalize_material(p, unit, title, date)
            if norm is None:
                food_list.append((unit, date, title, p))
            else:
                name = norm["name"]
                qty_num = extract_qty_number(p, name)
                col_key = "RUTAS" if "RUTAS" in unit else ("EXPLORADORES" if unit == "EXPLORADORES" else ("COMISIONES" if unit == "COMISIONES" else unit))
                
                consolidated[name]["detalles"].append(norm["literal"])
                if not consolidated[name]["category"]:
                    consolidated[name]["category"] = norm["category"]
                    consolidated[name]["type"] = norm["type"]
                
                # Asignar cantidad (si es dinámico, luego lo convertimos en fórmula)
                if consolidated[name][col_key] == 0:
                    consolidated[name][col_key] = qty_num
                else:
                    # Si ya tenía cantidad, la sumamos (ej. varios talleres o construcciones de la misma unidad)
                    if isinstance(consolidated[name][col_key], (int, float)):
                        consolidated[name][col_key] += qty_num

# Aplicar las fórmulas dinámicas de CONFIG
for name, info in consolidated.items():
    # 1. Hilo coreano Lobatos (3m por niño)
    if "hilo coreano" in name.lower():
        info["LOBATOS"] = "=CONFIG!$B$3*3"
    
    # 2. Cuentas Exploradores (30 por niño)
    if "cuentas" in name.lower() and "amarillas" in name.lower():
        info["EXPLORADORES"] = "=CONFIG!$B$4*30"
    if "cuentas" in name.lower() and "verde" in name.lower():
        info["EXPLORADORES"] = "=CONFIG!$B$4*30"
        
    # 3. Tote bags Pioneros (1 por chaval)
    if "tote bag" in name.lower() or "tote bags" in name.lower():
        info["PIONEROS"] = "=CONFIG!$B$5"
        
    # 4. Mandarinas Pioneros (1 por niño)
    if "mandarina" in name.lower():
        info["PIONEROS"] = "=CONFIG!$B$5"
        
    # 5. Bolsas de basura grande/mediana en Lobatos (1 por niño)
    if "bolsa" in name.lower() and "basura" in name.lower() and any("1 x niño" in d.lower() for d in info["detalles"]):
        info["LOBATOS"] = "=CONFIG!$B$3"
        
    # 6. Camisetas
    if "camiseta" in name.lower() and "blanca" in name.lower():
        if info["CASTORES"] > 0: info["CASTORES"] = "=CONFIG!$B$2"
        if info["LOBATOS"] > 0: info["LOBATOS"] = "=CONFIG!$B$3"
        if info["EXPLORADORES"] > 0: info["EXPLORADORES"] = "=CONFIG!$B$4"
        if info["PIONEROS"] > 0: info["PIONEROS"] = "=CONFIG!$B$5"
        
    # 7. Reglas de MASTER de data3.py (sobres, tubos PVC, cordino, arcilla)
    master_match = find_in_master(name.lower())
    if master_match and master_match[3] and master_match[3].startswith("="):
        # Mapear las fórmulas del MASTER
        f_raw = master_match[3]
        if "CONFIG!$B$3" in f_raw:
            info["LOBATOS"] = "=CONFIG!$B$3"
        if "CONFIG!$B$4" in f_raw:
            if "*30" in f_raw:
                info["EXPLORADORES"] = "=CONFIG!$B$4*30"
            else:
                info["EXPLORADORES"] = "=CONFIG!$B$4"
        if "CONFIG!$B$2" in f_raw:
            if "arcilla" in name.lower():
                info["CASTORES"] = "=CONFIG!$B$2*0.5" # 0.5kg arcilla por niño
            else:
                info["CASTORES"] = "=CONFIG!$B$2"
        if "CONFIG!$B$2" in f_raw and "CONFIG!$B$4" in f_raw:
            # Cordino: 1.5m por participante en Castores y Exploradores
            info["CASTORES"] = "=CONFIG!$B$2*1.5"
            info["EXPLORADORES"] = "=CONFIG!$B$4*1.5"

# ══════════════════════════════════════════════════════════════════════
# GENERACIÓN DE EXCEL
# ══════════════════════════════════════════════════════════════════════
print("Creando libro de trabajo openpyxl...")
wb = Workbook()

# Estilos premium
hdr_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
hdr_fill_navy = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
hdr_fill_green = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
hdr_fill_orange = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
hdr_fill_purple = PatternFill(start_color="5B315B", end_color="5B315B", fill_type="solid")

alt_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9")
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center")

def style_header_row(ws, row_idx, fill):
    ws.row_dimensions[row_idx].height = 28
    for cell in ws[row_idx]:
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def auto_fit_columns(ws, min_w=12, max_w=65):
    for col in ws.columns:
        max_len = min_w
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                lines = str(cell.value).split("\n")
                max_len = max(max_len, min(max(len(l) for l in lines) + 3, max_w))
        ws.column_dimensions[col_letter].width = max_len

# ── 1. CONFIG Sheet ───────────────────────────────────────────────────
ws_cfg = wb.active
ws_cfg.title = "CONFIG"
ws_cfg.append(["UNIDAD", "N DE NIÑOS"])
style_header_row(ws_cfg, 1, hdr_fill_purple)

for unit in ["CASTORES", "LOBATOS", "EXPLORADORES", "PIONEROS", "RUTAS PANDORA", "RUTAS ICARO"]:
    ws_cfg.append([unit, ""])
    rn = ws_cfg.max_row
    ws_cfg.cell(row=rn, column=1).font = Font(name="Segoe UI", bold=True)
    ws_cfg.cell(row=rn, column=2).alignment = center_align
    ws_cfg.cell(row=rn, column=1).border = thin_border
    ws_cfg.cell(row=rn, column=2).border = thin_border

ws_cfg.append([])
ws_cfg.append(["INSTRUCCIONES:", "Rellena la columna B con el número de niños de cada unidad para actualizar las cantidades dinámicas."])
ws_cfg.merge_cells("B10:G10")
ws_cfg.cell(row=10, column=2).font = Font(name="Segoe UI", italic=True, size=10, color="595959")
ws_cfg.column_dimensions["A"].width = 25
ws_cfg.column_dimensions["B"].width = 18

# ── 2. LISTA DE LA COMPRA Sheet (ESTRUCTURA DESGLOSADA) ─────────────────
ws_lista = wb.create_sheet(title="LISTA DE LA COMPRA", index=1)
headers = [
    "MATERIAL", "CATEGORÍA", "TOTAL", 
    "CASTORES", "LOBATOS", "EXPLORADORES (TROPA)", "PIONEROS", "RUTAS", "COMISIONES",
    "TIPO", "DETALLE (texto literal por actividad)", "CAJA", "COMPRADO ✓"
]
ws_lista.append(headers)
style_header_row(ws_lista, 1, hdr_fill_navy)

sorted_items = sorted(consolidated.items(), key=lambda x: x[0].lower())
for i, (material, info) in enumerate(sorted_items):
    detalles = "\n".join(info["detalles"])
    
    # Escribir la fila (con la fórmula SUM en la columna TOTAL)
    # TOTAL está en la columna 3 (C)
    # Las unidades están de la columna 4 (D) a la 8 (H)
    ws_lista.append([
        material, info["category"], "", # El total se insertará como fórmula en la siguiente línea
        info["CASTORES"], info["LOBATOS"], info["EXPLORADORES"], info["PIONEROS"], info["RUTAS"], info["COMISIONES"],
        info["type"], detalles, "", ""
    ])
    rn = ws_lista.max_row
    
    # Colocar la fórmula de SUMA en la columna TOTAL (celda C)
    ws_lista.cell(row=rn, column=3).value = f"=SUM(D{rn}:I{rn})"
    
    # Estilos de celdas
    ws_lista.cell(row=rn, column=1).font = Font(name="Segoe UI", bold=True)
    ws_lista.cell(row=rn, column=2).alignment = center_align
    ws_lista.cell(row=rn, column=3).alignment = center_align
    ws_lista.cell(row=rn, column=3).font = Font(name="Segoe UI", bold=True, color="1F497D")
    
    # Alinear al centro las cantidades de las unidades
    for col in range(4, 10):
        ws_lista.cell(row=rn, column=col).alignment = center_align
        
    ws_lista.cell(row=rn, column=10).alignment = center_align
    ws_lista.cell(row=rn, column=11).alignment = wrap_align
    
    for cell in ws_lista[rn]:
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = alt_fill

# Column widths para un diseño espacioso
col_widths = [35, 24, 12, 12, 12, 22, 12, 12, 12, 24, 70, 10, 12]
for idx, w in enumerate(col_widths, 1):
    ws_lista.column_dimensions[get_column_letter(idx)].width = w

# Validaciones de datos
dv_tipo = DataValidation(type="list",
    formula1='"COMPRAR,YA HAY (bases) - revisar,LO TRAE EL NINO,UNIDAD/COMISION,PEDIR A COCINA/MENU"',
    allow_blank=True)
ws_lista.add_data_validation(dv_tipo)
for row in range(2, ws_lista.max_row + 1):
    dv_tipo.add(ws_lista.cell(row=row, column=10)) # Columna J (10)

dv_cat = DataValidation(type="list",
    formula1='"Construccion/herramientas,Papeleria/manualidades,Juegos/deporte,Cocina/menaje,Iluminacion/electronica,Higiene/limpieza,Textil/disfraces,Impresiones/documentacion,Equipo campamento,Otros"',
    allow_blank=True)
ws_lista.add_data_validation(dv_cat)
for row in range(2, ws_lista.max_row + 1):
    dv_cat.add(ws_lista.cell(row=row, column=2)) # Columna B (2)

ws_lista.freeze_panes = "A2"
ws_lista.auto_filter.ref = ws_lista.dimensions

# ── 3. Unit sheets ──
for unit in UNITS_ORDER:
    ws = wb.create_sheet(title=unit[:31])
    ws.append(["FECHA / MOMENTO", "ACTIVIDAD", "MATERIAL (literal del docx)"])
    style_header_row(ws, 1, hdr_fill_green)
    
    activities = ALL_RAW_DATA.get(unit, [])
    for i, (date, title, material) in enumerate(activities):
        mat_clean = material.replace(" | ", "\n").replace("; ", "\n").replace(";", "\n")
        ws.append([date, title, mat_clean])
        rn = ws.max_row
        
        ws.cell(row=rn, column=1).alignment = wrap_align
        ws.cell(row=rn, column=2).font = Font(name="Segoe UI", bold=True)
        ws.cell(row=rn, column=2).alignment = wrap_align
        ws.cell(row=rn, column=3).alignment = wrap_align
        
        for cell in ws[rn]:
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = alt_fill
                
    auto_fit_columns(ws)
    ws.freeze_panes = "A2"

# ── 4. ALIMENTACIÓN Sheet ──
ws_food = wb.create_sheet(title="ALIMENTACIÓN")

ws_food.append(["-- INGREDIENTES DE TALLERES DE COCINA --"])
ws_food.merge_cells("A1:D1")
ws_food.cell(row=1, column=1).font = Font(name="Segoe UI", bold=True, size=13, color="C65911")
ws_food.row_dimensions[1].height = 25

ws_food.append(["UNIDAD", "FECHA / DÍA", "TALLER DE COCINA", "INGREDIENTES REQUERIDOS"])
style_header_row(ws_food, 2, hdr_fill_orange)

for i, (unit, date, title, ingredients) in enumerate(food_list):
    clean_ing = ingredients.replace("; ", "\n").replace(";", "\n")
    ws_food.append([unit, date, title, clean_ing])
    rn = ws_food.max_row
    
    ws_food.cell(row=rn, column=1).font = Font(name="Segoe UI", bold=True)
    ws_food.cell(row=rn, column=1).alignment = wrap_align
    ws_food.cell(row=rn, column=2).alignment = wrap_align
    ws_food.cell(row=rn, column=3).alignment = wrap_align
    ws_food.cell(row=rn, column=4).alignment = wrap_align
    
    for cell in ws_food[rn]:
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = alt_fill

ws_food.append([])
ws_food.append([])

ws_food.append(["-- MENÚ DE VOLANTES (Ingredientes Desglosados) --"])
ws_food.merge_cells(start_row=ws_food.max_row, start_column=1, end_row=ws_food.max_row, end_column=4)
ws_food.cell(row=ws_food.max_row, column=1).font = Font(name="Segoe UI", bold=True, size=13, color="C65911")
ws_food.row_dimensions[ws_food.max_row].height = 25

ws_food.append(["UNIDAD", "FECHA / DÍA", "ACTIVIDAD / ORIGEN", "INGREDIENTE / CANTIDAD"])
style_header_row(ws_food, ws_food.max_row, hdr_fill_orange)

for i, (unit, date, title, item) in enumerate(menu_items):
    is_cleaning = any(c in item.lower() for c in ["fairy", "estropajo", "basura", "albal", "chalecos", "higiénico", "botiquín", "walkie"])
    dest_item = item + " (limpieza/material - ver lista principal)" if is_cleaning else item
    
    ws_food.append([unit, date, title, dest_item])
    rn = ws_food.max_row
    
    ws_food.cell(row=rn, column=1).font = Font(name="Segoe UI", bold=True)
    ws_food.cell(row=rn, column=1).alignment = wrap_align
    ws_food.cell(row=rn, column=2).alignment = wrap_align
    ws_food.cell(row=rn, column=3).alignment = wrap_align
    ws_food.cell(row=rn, column=4).alignment = wrap_align
    
    for cell in ws_food[rn]:
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = alt_fill

auto_fit_columns(ws_food)
ws_food.freeze_panes = "A2"

# ── 5. REPARTO CAJAS Sheet ──
ws_cajas = wb.create_sheet(title="REPARTO CAJAS")
ws_cajas.append(["MATERIAL", "CAJA", "CANTIDAD REPARTIDA"])
style_header_row(ws_cajas, 1, hdr_fill_navy)
ws_cajas.column_dimensions["A"].width = 35
ws_cajas.column_dimensions["B"].width = 15
ws_cajas.column_dimensions["C"].width = 20
ws_cajas.freeze_panes = "A2"

# Guardar Excel
output_path = os.path.join(BASE, "LISTA_COMPRA_CAMPAMENTO.xlsx")
wb.save(output_path)

print(f"\nExcel desglosado generado con éxito en: {output_path}")
print("  - Estructura: 5 columnas desglosadas por unidad + TOTAL dinámico con fórmulas SUM.")
